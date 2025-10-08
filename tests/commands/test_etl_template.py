# these tests can be run from the command line via
# python manage.py test tests.commands.test_etl_template --settings="tests.test_settings"

import os

from django.core.management import call_command
from openpyxl import load_workbook

from arches.app.models.models import CardModel
from tests.base_test import ArchesTestCase


class ETLTemplateTests(ArchesTestCase):
    graph_fixtures = ["Data_Type_Model"]

    def test_sheet_title_replaces_slashes(self):
        card = CardModel.objects.get(name__en__iexact="Images-Files")
        card.name = "Images / Files"
        card.save()

        dest = "test_template.xlsx"
        self.addCleanup(os.remove, dest)
        call_command(
            "etl_template", template="tilexls", graph=str(card.graph.pk), dest=dest
        )

        wb = load_workbook(dest, read_only=True)
        self.assertIn("Images _ Files", wb.sheetnames)
